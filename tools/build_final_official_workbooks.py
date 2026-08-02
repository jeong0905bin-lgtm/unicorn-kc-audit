from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
DIAG = ROOT / "diagnostics"
OUT = ROOT / "final-output"
EXCLUDED_KC = "U003E1577-7011"


def load_json(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def normalize_product(row: dict[str, Any]) -> dict[str, Any]:
    product_id = str(row.get("productId") or "")
    item_id = str(row.get("itemId") or row.get("catalogItemId") or "")
    name = str(row.get("productName") or "")
    unique_id = str(row.get("coupangUniqueId") or "").strip()
    if not unique_id and product_id and item_id:
        unique_id = f"{product_id} - {item_id}"
    kc_numbers = row.get("kcNumbers") or []
    if row.get("kcNumber"):
        kc_numbers = list(kc_numbers) + [row["kcNumber"]]
    kc_numbers = sorted({str(v).strip() for v in kc_numbers if str(v).strip() and str(v).strip() != EXCLUDED_KC})
    return {
        "productId": product_id,
        "itemId": item_id,
        "productName": name,
        "coupangUniqueId": unique_id,
        "publisher": str(row.get("publisher") or "유니콘"),
        "publisherEvidenceType": str(row.get("publisherEvidenceType") or row.get("evidenceType") or ""),
        "kcNumbers": kc_numbers,
        "kcAbsenceConfirmed": row.get("kcAbsenceConfirmed") is True,
    }


def merge_product(target: dict[str, dict[str, Any]], row: dict[str, Any]) -> None:
    p = normalize_product(row)
    key = p["productId"] or p["coupangUniqueId"]
    if not key:
        return
    current = target.get(key)
    if current is None:
        target[key] = p
        return
    current["kcNumbers"] = sorted(set(current["kcNumbers"]) | set(p["kcNumbers"]))
    current["kcAbsenceConfirmed"] = bool(current.get("kcAbsenceConfirmed") or p.get("kcAbsenceConfirmed"))
    for field in ("itemId", "productName", "coupangUniqueId", "publisherEvidenceType"):
        if not current.get(field) and p.get(field):
            current[field] = p[field]


def collect_products() -> dict[str, dict[str, Any]]:
    products: dict[str, dict[str, Any]] = {}

    consolidated = load_json(DIAG / "consolidated-unicorn-products.json")
    for row in consolidated.get("products", []):
        merge_product(products, row)

    verified = load_json(DIAG / "verified-unicorn-products.json")
    for row in verified.get("verifiedProducts", []):
        merge_product(products, row)

    for filename in ("kc-mapping-addendum-20260801.json", "kc-mapping-addendum-20260801-run1.json"):
        data = load_json(DIAG / filename)
        for row in data.get("newConfirmedMappings", []):
            merge_product(products, row)

    official = load_json(DIAG / "official-status-recovery-20260802-run1.json")
    for row in official.get("expiredProductRowsConfirmedFromExistingExactMappings", []):
        merge_product(products, row)
    for row in official.get("activeMappedProductsExcludedFromExpiredOutput", []):
        merge_product(products, row)

    priority = load_json(DIAG / "unresolved-priority-resolution-20260802-run14.json")
    for bucket in ("publisherConfirmedExact", "publisherPromotedInferredHigh"):
        for row in priority.get(bucket, []):
            merge_product(products, row)

    recovery = load_json(DIAG / "publisher-kc-recovery-20260802-run2.json")
    for bucket in ("newExactPublisherConfirmations", "newExactKcMappings"):
        for row in recovery.get(bucket, []):
            merge_product(products, row)

    latest = load_json(DIAG / "official-audit-addendum-20260802-run16.json")
    for bucket in ("newExactPublisherConfirmations", "newExactKcMappings", "kcAbsenceConfirmedProducts"):
        for row in latest.get(bucket, []):
            merge_product(products, row)

    return products


def collect_statuses() -> dict[str, dict[str, Any]]:
    statuses: dict[str, dict[str, Any]] = {}
    official = load_json(DIAG / "official-status-recovery-20260802-run1.json")
    for kc, info in official.get("officialSafetyKoreaStatuses", {}).items():
        if kc != EXCLUDED_KC:
            statuses[kc] = dict(info)
    state = load_json(DIAG / "consolidated-unicorn-products.json")
    for kc, info in state.get("officialSafetyKorea", {}).items():
        if kc != EXCLUDED_KC and kc not in statuses:
            statuses[kc] = dict(info)
    latest = load_json(DIAG / "official-audit-addendum-20260802-run16.json")
    for kc, info in latest.get("officialSafetyKoreaStatuses", {}).items():
        if kc != EXCLUDED_KC:
            statuses[kc] = dict(info)
    return statuses


def style_header(ws, row: int, cols: int, fill: str) -> None:
    for cell in ws[row][:cols]:
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")


def save_no_kc(products: list[dict[str, Any]]) -> int:
    rows = [p for p in products if p.get("kcAbsenceConfirmed") is True]
    wb = Workbook()
    ws = wb.active
    ws.title = "No-KC"
    ws.append(["상품명", "쿠팡 상품 고유번호", "KC 인증번호"])
    for p in rows:
        ws.append([p["productName"], p["coupangUniqueId"], "없음"])
    style_header(ws, 1, 3, "1F4E78")
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 18
    meta = wb.create_sheet("검증정보")
    meta.append(["항목", "내용"])
    meta.append(["검증 완료 No-KC 행 수", len(rows)])
    meta.append(["원칙", "차단·빈 응답·KC 미매핑은 없음으로 처리하지 않음"])
    meta.append(["제외 인증번호", EXCLUDED_KC])
    style_header(meta, 1, 2, "5B9BD5")
    meta.column_dimensions["A"].width = 28
    meta.column_dimensions["B"].width = 90
    wb.save(OUT / "A00214628_유니콘_No-KC_최종.xlsx")
    return len(rows)


def save_expired(products: list[dict[str, Any]], statuses: dict[str, dict[str, Any]]) -> int:
    rows: list[list[str]] = []
    for p in products:
        for kc in p.get("kcNumbers", []):
            info = statuses.get(kc, {})
            if info.get("officialExactMatch") is True and (info.get("expired") is True or info.get("status") == "기간만료"):
                rows.append([kc, p["productName"], p["coupangUniqueId"]])
    unique = []
    seen = set()
    for row in rows:
        key = tuple(row)
        if key not in seen:
            seen.add(key)
            unique.append(row)
    unique.sort(key=lambda r: (r[0], r[2], r[1]))

    wb = Workbook()
    ws = wb.active
    ws.title = "기간만료 KC"
    ws.append(["기한만료 KC 인증번호", "만료된 상품명", "쿠팡 상품 고유번호"])
    for row in unique:
        ws.append(row)
    style_header(ws, 1, 3, "C00000")
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 28

    evidence = wb.create_sheet("공식 상태")
    evidence.append(["KC 인증번호", "SafetyKorea 상태", "제조사", "모델명", "공식 exact"])
    for kc, info in sorted(statuses.items()):
        evidence.append([kc, info.get("status", "미확정"), info.get("manufacturer", ""), info.get("modelName", ""), info.get("officialExactMatch", False)])
    style_header(evidence, 1, 5, "5B9BD5")
    for col, width in zip("ABCDE", (26, 20, 26, 26, 18)):
        evidence.column_dimensions[col].width = width

    unresolved = wb.create_sheet("미해결 상품")
    unresolved.append(["상품명", "쿠팡 상품 고유번호", "사유"])
    for p in products:
        if not p.get("kcNumbers") and not p.get("kcAbsenceConfirmed"):
            unresolved.append([p["productName"], p["coupangUniqueId"], "상품별 exact KC 미매핑"])
        else:
            unknown = [kc for kc in p["kcNumbers"] if not statuses.get(kc, {}).get("officialExactMatch")]
            if unknown:
                unresolved.append([p["productName"], p["coupangUniqueId"], "공식 상태 미확정: " + ", ".join(unknown)])
    style_header(unresolved, 1, 3, "7F6000")
    unresolved.column_dimensions["A"].width = 52
    unresolved.column_dimensions["B"].width = 28
    unresolved.column_dimensions["C"].width = 48

    wb.save(OUT / "A00214628_유니콘_기간만료_KC_최종.xlsx")
    return len(unique)


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    product_map = collect_products()
    products = sorted(product_map.values(), key=lambda p: (int(p["productId"]) if p["productId"].isdigit() else 10**30, p["itemId"]))
    statuses = collect_statuses()
    no_kc_count = save_no_kc(products)
    expired_count = save_expired(products, statuses)
    summary = {
        "acceptedProducts": len(products),
        "productsWithExactKc": sum(bool(p["kcNumbers"]) for p in products),
        "uniqueKcNumbers": sorted({kc for p in products for kc in p["kcNumbers"]}),
        "officialStatuses": statuses,
        "noKcRows": no_kc_count,
        "expiredRows": expired_count,
        "excludedCertificationNumber": EXCLUDED_KC,
    }
    (OUT / "summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
