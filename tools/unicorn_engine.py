#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import html
import json
import re
import time
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from io import BytesIO
from pathlib import Path
from urllib.parse import parse_qs, urlencode, urlparse, urlsplit, urlunsplit

import cv2
import numpy as np
import pytesseract
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image
from playwright.sync_api import sync_playwright

UA_DESKTOP = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/149.0.0.0 Safari/537.36"
UA_MOBILE = "Mozilla/5.0 (Linux; Android 15; SM-S928N) AppleWebKit/537.36 Chrome/149.0.0.0 Mobile Safari/537.36"
BLOCK_TERMS = ("access denied", "captcha", "보안 확인", "로그인이 필요", "접근 불가", "비정상적인 접근")
KC_RE = re.compile(r"\b(?:[A-Z]{1,3}\d{2,4}[A-Z]\d{3,4}-\d{4}[A-Z]?)\b")
LOOSE_KC_RE = re.compile(r"\b([A-Z]{1,3}[0-9OIL]{2,4}[A-Z][0-9OIL]{3,4}-[0-9OIL]{4}[A-Z]?)\b")
SAFETY_URL = "https://www.safetykorea.kr/search/searchPop"


def save_json(path: Path, data: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp.replace(path)


def norm_space(value: str) -> str:
    return re.sub(r"\s+", " ", html.unescape(value or "")).strip()


def norm_value(value: str) -> str:
    return re.sub(r"^[\s:：|,·ㆍ/\-]+|[\s:：|,·ㆍ/\-]+$", "", norm_space(value)).strip()


def unique_id(product: dict) -> str:
    product_id = str(product.get("productId") or "")
    item_id = str(product.get("itemId") or "")
    return f"{product_id} - {item_id}" if product_id and item_id else product_id


def product_urls(product: dict) -> list[str]:
    product_id = str(product.get("productId") or "")
    item_id = str(product.get("itemId") or "")
    vendor_item_id = str(product.get("vendorItemId") or "")
    query = []
    if item_id:
        query.append("itemId=" + item_id)
    if vendor_item_id:
        query.append("vendorItemId=" + vendor_item_id)
    suffix = "?" + "&".join(query) if query else ""
    urls = [str(product.get("productUrl") or "")]
    if product_id:
        urls.extend([
            f"https://www.coupang.com/vp/products/{product_id}{suffix}",
            f"https://m.coupang.com/vm/products/{product_id}{suffix}",
        ])
    return list(dict.fromkeys(url for url in urls if url))


def cache_bust(url: str, attempt: int) -> str:
    parts = urlsplit(url)
    query = dict(parse_qs(parts.query, keep_blank_values=True))
    query = {k: v[-1] if isinstance(v, list) else v for k, v in query.items()}
    query["_fresh"] = f"{int(time.time() * 1000)}-{attempt}"
    return urlunsplit((parts.scheme, parts.netloc, parts.path, urlencode(query), parts.fragment))


def normalize_cdn(raw: str) -> str:
    raw = html.unescape(raw or "").replace("\\/", "/").replace("&amp;", "&")
    if raw.startswith("//"):
        raw = "https:" + raw
    if not raw.startswith("http") or "coupangcdn.com" not in raw:
        return ""
    match = re.search(r"/thumbnails/remote/(?:[^/]+/)?image/(.+)$", raw)
    if match:
        return "https://image1.coupangcdn.com/image/" + match.group(1).split("?")[0]
    return raw.split("?")[0]


def extract_images(text: str, soup: BeautifulSoup | None = None) -> list[str]:
    found = set()
    clean = html.unescape(text or "").replace("\\/", "/")
    for raw in re.findall(r'(?:https?:)?//[^\s"\'<>]+coupangcdn\.com/[^\s"\'<>]+', clean, re.I):
        url = normalize_cdn(raw)
        if url:
            found.add(url)
    if soup is not None:
        for tag in soup.find_all("img"):
            for key in ("src", "data-src", "data-original", "data-url"):
                url = normalize_cdn(tag.get(key) or "")
                if url:
                    found.add(url)
    return sorted(found)


def extract_title(soup: BeautifulSoup, fallback: str) -> str:
    for selector, attr in (("meta[property='og:title']", "content"), ("h1", None), ("title", None)):
        node = soup.select_one(selector)
        if node:
            value = node.get(attr, "") if attr else node.get_text(" ", strip=True)
            value = re.sub(r"\s*[-|]\s*쿠팡.*$", "", norm_space(value)).strip()
            if value:
                return value[:500]
    return norm_space(fallback)[:500]


def extract_publisher(text: str, soup: BeautifulSoup | None = None) -> tuple[str, str]:
    if soup is not None:
        for row in soup.select("tr"):
            cells = row.find_all(["th", "td"])
            if len(cells) >= 2:
                key = re.sub(r"\s+", "", cells[0].get_text(" ", strip=True))
                if "저자" in key and "출판사" in key:
                    return norm_value(cells[1].get_text(" ", strip=True)), "table"
        for dl in soup.select("dl"):
            for dt in dl.find_all("dt"):
                key = re.sub(r"\s+", "", dt.get_text(" ", strip=True))
                if "저자" in key and "출판사" in key:
                    dd = dt.find_next_sibling("dd")
                    if dd:
                        return norm_value(dd.get_text(" ", strip=True)), "dl"
    clean = html.unescape(text or "")
    patterns = [
        r"저자\s*[,·/ㆍ]\s*출판사\s*[:：]?\s*([^\n|<>]{1,100})",
        r"저자\s*출판사\s*[:：]?\s*([^\n|<>]{1,100})",
    ]
    for pattern in patterns:
        match = re.search(pattern, clean, re.I)
        if match:
            value = norm_value(re.split(r"배송|교환|반품|크기|쪽수", match.group(1))[0])
            if value:
                return value[:100], "regex"
    return "", "missing"


def normalize_kc(code: str) -> str:
    code = re.sub(r"\s+", "", (code or "").upper())
    match = re.fullmatch(r"([A-Z]{1,3})([0-9OIL]{2,4})([A-Z])([0-9OIL]{3,4})-([0-9OIL]{4})([A-Z]?)", code)
    if not match:
        return ""
    trans = str.maketrans({"O": "0", "I": "1", "L": "1"})
    fixed = match.group(1) + match.group(2).translate(trans) + match.group(3) + match.group(4).translate(trans) + "-" + match.group(5).translate(trans) + match.group(6)
    return fixed if KC_RE.fullmatch(fixed) else ""


def extract_kc(text: str) -> set[str]:
    compact = re.sub(r"\s+", "", (text or "").upper())
    found = set(KC_RE.findall(compact))
    for candidate in LOOSE_KC_RE.findall(compact):
        code = normalize_kc(candidate)
        if code:
            found.add(code)
    return found


def direct_detail(product: dict, attempt: int) -> dict | None:
    session = requests.Session()
    headers = {
        "User-Agent": UA_MOBILE if attempt % 2 == 0 else UA_DESKTOP,
        "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.7",
        "Cache-Control": "no-cache, no-store, max-age=0",
        "Pragma": "no-cache",
    }
    try:
        for url in product_urls(product):
            try:
                response = session.get(cache_bust(url, attempt), headers=headers, timeout=12)
            except requests.RequestException:
                continue
            if response.status_code != 200 or len(response.content) < 1500:
                continue
            response.encoding = response.apparent_encoding or "utf-8"
            raw = response.text
            if any(term in raw.lower() for term in BLOCK_TERMS):
                continue
            soup = BeautifulSoup(raw, "html.parser")
            text = soup.get_text("\n", strip=True)
            publisher, source = extract_publisher(text + "\n" + raw, soup)
            return {
                "source": "requests",
                "productName": extract_title(soup, product.get("sourceName") or ""),
                "publisherValue": publisher,
                "publisherSource": source,
                "imageUrls": extract_images(raw, soup),
                "pageKcNumbers": sorted(extract_kc(text + "\n" + raw)),
                "accessConfirmed": True,
                "finalUrl": response.url,
                "error": "",
            }
    finally:
        session.close()
    return None


def browser_detail(browser, product: dict, attempt: int) -> dict | None:
    mobile = attempt % 2 == 0
    context = browser.new_context(
        user_agent=UA_MOBILE if mobile else UA_DESKTOP,
        locale="ko-KR",
        timezone_id="Asia/Seoul",
        viewport={"width": 430, "height": 932} if mobile else {"width": 1440, "height": 1200},
        is_mobile=mobile,
        has_touch=mobile,
        service_workers="block",
    )
    context.set_default_timeout(12000)
    page = context.new_page()
    network = []
    images = set()

    def on_response(response):
        try:
            ctype = (response.headers.get("content-type") or "").lower()
            if "image" in ctype:
                url = normalize_cdn(response.url)
                if url:
                    images.add(url)
            elif any(x in ctype for x in ("json", "html", "text", "javascript")):
                body = response.text()
                if len(body) <= 10_000_000:
                    network.append(body)
                    images.update(extract_images(body))
        except Exception:
            pass

    page.on("response", on_response)
    try:
        for url in product_urls(product):
            try:
                page.goto(cache_bust(url, attempt), wait_until="domcontentloaded", timeout=18000)
                page.wait_for_timeout(900)
                page.evaluate("window.scrollBy(0, Math.max(1000, window.innerHeight)); window.dispatchEvent(new Event('scroll'))")
                page.wait_for_timeout(500)
                body = page.locator("body").inner_text(timeout=5000)
                content = page.content()
                if any(term in body.lower() for term in BLOCK_TERMS):
                    continue
                combined = body + "\n" + content + "\n" + "\n".join(network)
                soup = BeautifulSoup(content, "html.parser")
                publisher, source = extract_publisher(combined, soup)
                images.update(extract_images(combined, soup))
                return {
                    "source": "playwright",
                    "productName": extract_title(soup, product.get("sourceName") or ""),
                    "publisherValue": publisher,
                    "publisherSource": source,
                    "imageUrls": sorted(images),
                    "pageKcNumbers": sorted(extract_kc(combined)),
                    "accessConfirmed": True,
                    "finalUrl": page.url,
                    "error": "",
                }
            except Exception:
                continue
    finally:
        context.close()
    return None


def run_filter(args) -> None:
    catalog = json.loads(args.catalog.read_text(encoding="utf-8"))
    products = [p for i, p in enumerate(catalog.get("products", [])) if i % args.shards == args.shard]
    results = []
    browser = None
    pw = None
    try:
        for index, product in enumerate(products, 1):
            detail = direct_detail(product, 1) or direct_detail(product, 2)
            if detail is None or not detail.get("publisherValue"):
                if pw is None:
                    pw = sync_playwright().start()
                    browser = pw.chromium.launch(headless=True, args=["--disable-dev-shm-usage", "--no-sandbox"])
                fallback = browser_detail(browser, product, 1) or browser_detail(browser, product, 2)
                if fallback and (detail is None or fallback.get("publisherValue") or len(fallback.get("imageUrls", [])) > len(detail.get("imageUrls", []))):
                    detail = fallback
            detail = detail or {
                "source": "none", "productName": product.get("sourceName") or "", "publisherValue": "",
                "publisherSource": "missing", "imageUrls": [], "pageKcNumbers": [], "accessConfirmed": False,
                "finalUrl": product.get("productUrl") or "", "error": "inaccessible",
            }
            compact = re.sub(r"\s+", "", norm_value(detail.get("publisherValue") or ""))
            status = "unicorn" if compact == "유니콘" else ("other" if compact else ("missing" if detail.get("accessConfirmed") else "inaccessible"))
            row = {**product, **detail, "publisherStatus": status, "coupangUniqueId": unique_id(product)}
            results.append(row)
            save_json(args.output.with_suffix(".checkpoint.json"), {"shard": args.shard, "processed": index, "total": len(products), "products": results})
            print(json.dumps({"phase": "filter", "shard": args.shard, "processed": index, "total": len(products), "status": status}, ensure_ascii=False), flush=True)
    finally:
        if browser is not None:
            browser.close()
        if pw is not None:
            pw.stop()
    save_json(args.output, {"shard": args.shard, "count": len(results), "products": results})


def image_variants(raw: bytes) -> list[np.ndarray]:
    image = Image.open(BytesIO(raw)).convert("RGB")
    arr = cv2.cvtColor(np.array(image), cv2.COLOR_RGB2BGR)
    if arr.shape[1] < 1100:
        scale = min(2.4, 1100 / max(1, arr.shape[1]))
        arr = cv2.resize(arr, None, fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)
    gray = cv2.cvtColor(arr, cv2.COLOR_BGR2GRAY)
    variants = [gray, cv2.equalizeHist(gray), cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1]]
    return variants


def ocr_bytes(raw: bytes) -> tuple[Counter, str]:
    votes = Counter()
    texts = []
    try:
        variants = image_variants(raw)
    except Exception as exc:
        return votes, f"decode:{exc}"
    for variant in variants:
        height = variant.shape[0]
        tiles = [variant] if height <= 5000 else [variant[y:min(y + 4500, height), :] for y in range(0, height, 4000)]
        for tile in tiles:
            for psm in (6, 11):
                try:
                    text = pytesseract.image_to_string(tile, lang="kor+eng", config=f"--psm {psm}")
                except Exception:
                    text = pytesseract.image_to_string(tile, lang="eng", config=f"--psm {psm}")
                texts.append(text)
                for code in extract_kc(text):
                    votes[code] += 1
    return votes, "\n".join(texts)


def run_ocr(args) -> None:
    data = json.loads(args.input.read_text(encoding="utf-8"))
    products = [p for i, p in enumerate(data.get("products", [])) if i % args.shards == args.shard]
    results = []
    session = requests.Session()
    session.headers.update({"User-Agent": UA_DESKTOP, "Accept-Language": "ko-KR,ko;q=0.9"})
    args.image_dir.mkdir(parents=True, exist_ok=True)
    for index, product in enumerate(products, 1):
        page_codes = set(product.get("pageKcNumbers") or [])
        votes = Counter({code: 99 for code in page_codes})
        downloaded = 0
        processed = 0
        errors = []
        urls = list(dict.fromkeys(product.get("imageUrls") or []))[:40]
        for image_no, url in enumerate(urls, 1):
            try:
                response = session.get(url, timeout=18)
                ctype = (response.headers.get("content-type") or "").lower()
                if response.status_code != 200 or "image" not in ctype or not (500 <= len(response.content) <= 18_000_000):
                    continue
                downloaded += 1
                image_votes, _ = ocr_bytes(response.content)
                processed += 1
                votes.update(image_votes)
            except Exception as exc:
                errors.append(f"{image_no}:{type(exc).__name__}")
        accepted = sorted(code for code, count in votes.items() if count >= 2 or code in page_codes)
        review = sorted(code for code, count in votes.items() if count == 1 and code not in accepted)
        if accepted:
            status = "found"
        elif review:
            status = "candidate_review"
        elif downloaded > 0 and processed > 0:
            status = "no_kc"
        elif urls:
            status = "image_download_failed"
        else:
            status = "no_images"
        row = {**product, "resultStatus": status, "kcNumbers": accepted, "kcCandidates": review, "imageUrlsAttempted": len(urls), "imagesDownloaded": downloaded, "imagesProcessed": processed, "ocrVoteCounts": dict(votes), "ocrErrors": errors[:30]}
        results.append(row)
        save_json(args.output.with_suffix(".checkpoint.json"), {"shard": args.shard, "processed": index, "total": len(products), "products": results})
        print(json.dumps({"phase": "ocr", "shard": args.shard, "processed": index, "total": len(products), "status": status, "kc": accepted}, ensure_ascii=False), flush=True)
    session.close()
    save_json(args.output, {"shard": args.shard, "count": len(results), "products": results})


def find_value(soup: BeautifulSoup, label: str) -> str:
    wanted = norm_space(label)
    for node in soup.find_all(["th", "td", "dt", "dd", "span", "strong", "div"]):
        if norm_space(node.get_text(" ", strip=True)) != wanted:
            continue
        sibling = node.find_next_sibling()
        if sibling:
            value = norm_space(sibling.get_text(" ", strip=True))
            if value:
                return value
        parent = node.parent
        if parent:
            cells = parent.find_all(["th", "td", "dt", "dd"], recursive=False)
            for i, cell in enumerate(cells):
                if cell is node and i + 1 < len(cells):
                    return norm_space(cells[i + 1].get_text(" ", strip=True))
    return ""


def lookup_one(code: str) -> dict:
    last_error = ""
    for attempt in range(1, 5):
        try:
            response = requests.get(
                SAFETY_URL,
                params={"certNum": code, "menu": "search", "_fresh": f"{int(time.time() * 1000)}-{attempt}"},
                headers={"User-Agent": UA_DESKTOP, "Accept-Language": "ko-KR,ko;q=0.9", "Referer": "https://www.safetykorea.kr/release/itemSearch"},
                timeout=30,
            )
            if response.status_code != 200:
                raise RuntimeError(f"HTTP {response.status_code}")
            soup = BeautifulSoup(response.text, "html.parser")
            official = find_value(soup, "인증번호")
            status = find_value(soup, "인증상태")
            exact = re.sub(r"\s+", "", official.upper()) == code.upper()
            return {"kcNumber": code, "officialNumber": official, "officialStatus": status, "officialExactMatch": exact, "isExpired": exact and any(word in status for word in ("기간만료", "기한만료")), "resultStatus": "exact" if exact else ("not_found" if not official and not status else "mismatch"), "officialUrl": response.url}
        except Exception as exc:
            last_error = f"{type(exc).__name__}: {exc}"
            time.sleep(attempt)
    return {"kcNumber": code, "officialNumber": "", "officialStatus": "", "officialExactMatch": False, "isExpired": False, "resultStatus": "error", "error": last_error, "officialUrl": f"{SAFETY_URL}?certNum={code}&menu=search"}


def write_csv(path: Path, headers: list[str], rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(path: Path, sheet: str, headers: list[str], rows: list[dict], widths: list[int]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + index)].width = width
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def run_lookup_build(args) -> None:
    data = json.loads(args.collection.read_text(encoding="utf-8"))
    products = data.get("products", [])
    codes = sorted({code for product in products for code in (product.get("kcNumbers") or [])})
    results = []
    with ThreadPoolExecutor(max_workers=max(1, args.workers)) as pool:
        futures = {pool.submit(lookup_one, code): code for code in codes}
        for future in as_completed(futures):
            result = future.result()
            results.append(result)
            print(json.dumps({"phase": "lookup", "kc": result["kcNumber"], "status": result.get("officialStatus"), "result": result.get("resultStatus")}, ensure_ascii=False), flush=True)
    results.sort(key=lambda item: item["kcNumber"])
    status_map = {item["kcNumber"]: item for item in results}
    no_kc_rows = []
    expired_rows = []
    review_rows = []
    for product in products:
        name = product.get("productName") or product.get("sourceName") or ""
        uid = product.get("coupangUniqueId") or unique_id(product)
        if product.get("resultStatus") == "no_kc":
            no_kc_rows.append({"상품명": name, "쿠팡 상품 고유번호": uid, "KC 인증번호": "없음"})
        for code in product.get("kcNumbers") or []:
            official = status_map.get(code) or {}
            if official.get("isExpired"):
                expired_rows.append({"기한만료 KC 인증번호": code, "만료된 상품명": name, "쿠팡 상품 고유번호": uid})
        if product.get("resultStatus") in {"candidate_review", "image_download_failed", "no_images", "error"}:
            review_rows.append({"상품명": name, "쿠팡 상품 고유번호": uid, "검토상태": product.get("resultStatus"), "KC 후보": " | ".join(product.get("kcCandidates") or [])})
    no_kc_rows = sorted({(r["상품명"], r["쿠팡 상품 고유번호"]): r for r in no_kc_rows}.values(), key=lambda r: (r["상품명"], r["쿠팡 상품 고유번호"]))
    expired_rows = sorted({(r["기한만료 KC 인증번호"], r["만료된 상품명"], r["쿠팡 상품 고유번호"]): r for r in expired_rows}.values(), key=lambda r: (r["기한만료 KC 인증번호"], r["만료된 상품명"], r["쿠팡 상품 고유번호"]))
    out = args.output_dir
    out.mkdir(parents=True, exist_ok=True)
    no_headers = ["상품명", "쿠팡 상품 고유번호", "KC 인증번호"]
    expired_headers = ["기한만료 KC 인증번호", "만료된 상품명", "쿠팡 상품 고유번호"]
    write_csv(out / "kc_no_number_products.csv", no_headers, no_kc_rows)
    write_csv(out / "expired_kc_products.csv", expired_headers, expired_rows)
    write_csv(out / "manual_review_products.csv", ["상품명", "쿠팡 상품 고유번호", "검토상태", "KC 후보"], review_rows)
    write_xlsx(out / "유니콘_KC인증번호_없는상품.xlsx", "KC 번호 없음", no_headers, no_kc_rows, [70, 28, 18])
    write_xlsx(out / "유니콘_기간만료_KC상품.xlsx", "기간만료 KC", expired_headers, expired_rows, [25, 70, 28])
    save_json(out / "kc_official_status.json", {"count": len(results), "results": results})
    save_json(out / "kc_no_number_products.json", {"count": len(no_kc_rows), "products": no_kc_rows})
    save_json(out / "expired_kc_products.json", {"count": len(expired_rows), "products": expired_rows})
    save_json(out / "manual_review_products.json", {"count": len(review_rows), "products": review_rows})
    summary = {
        "unicornProducts": len(products),
        "productsWithKc": sum(bool(p.get("kcNumbers")) for p in products),
        "uniqueKcNumbers": len(codes),
        "noKcProducts": len(no_kc_rows),
        "expiredRows": len(expired_rows),
        "manualReview": len(review_rows),
        "statusCounts": dict(Counter(str(p.get("resultStatus") or "unknown") for p in products)),
    }
    save_json(out / "summary.json", summary)
    print(json.dumps(summary, ensure_ascii=False), flush=True)


def main() -> None:
    parser = argparse.ArgumentParser()
    sub = parser.add_subparsers(dest="command", required=True)
    p_filter = sub.add_parser("filter")
    p_filter.add_argument("--catalog", type=Path, required=True)
    p_filter.add_argument("--output", type=Path, required=True)
    p_filter.add_argument("--shard", type=int, required=True)
    p_filter.add_argument("--shards", type=int, required=True)
    p_filter.set_defaults(func=run_filter)
    p_ocr = sub.add_parser("ocr")
    p_ocr.add_argument("--input", type=Path, required=True)
    p_ocr.add_argument("--output", type=Path, required=True)
    p_ocr.add_argument("--image-dir", type=Path, required=True)
    p_ocr.add_argument("--shard", type=int, required=True)
    p_ocr.add_argument("--shards", type=int, required=True)
    p_ocr.set_defaults(func=run_ocr)
    p_build = sub.add_parser("lookup-build")
    p_build.add_argument("--collection", type=Path, required=True)
    p_build.add_argument("--output-dir", type=Path, required=True)
    p_build.add_argument("--workers", type=int, default=10)
    p_build.set_defaults(func=run_lookup_build)
    args = parser.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
