#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import quote_plus

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
UNICORN_RE = re.compile(r"(?:\(주\)\s*유니콘|주식회사\s*유니콘|BOOKFRIENDS|UNICORN|(?<![가-힣])유니콘(?![가-힣]))", re.I)
CANDIDATE_RE = re.compile(r"편집부|출판사\s*미기재|제조사\s*미기재|스티커|색칠북|판퍼즐|대판퍼즐|퍼즐|워터색칠북", re.I)
KC_RE = re.compile(r"\b(?:CB|CA|SU|U)\d{3,}[A-Z0-9-]{5,}\b", re.I)
EXCLUDED_KC = {"U003E1577-7011"}
HEADERS = {"User-Agent": "Mozilla/5.0", "Accept-Language": "ko-KR,ko;q=0.9"}


def now():
    return datetime.now(timezone.utc).isoformat()


def read_shards(path: Path):
    merged = {}
    source_states = []
    for file in sorted(path.glob("*.json")):
        data = json.loads(file.read_text(encoding="utf-8"))
        source_states.extend(data.get("sourceStates", []))
        for row in data.get("products", []):
            pid = str(row.get("productId", ""))
            if not pid:
                continue
            current = merged.setdefault(pid, row)
            for k, v in row.items():
                if (not current.get(k)) and v:
                    current[k] = v
            current["sourceUrls"] = sorted(set(current.get("sourceUrls", []) + row.get("sourceUrls", [])))
            if row.get("responseState") == "ok":
                current["responseState"] = "ok"
    return list(merged.values()), source_states


def classify(row):
    text = " ".join(str(row.get(k, "")) for k in ["productName", "publisherManufacturer", "brand", "kcText"])
    if UNICORN_RE.search(text):
        return "확정", "쿠팡 직접표기"
    if CANDIDATE_RE.search(text):
        return "후보", "상품군 확장검색"
    return "제외", "유니콘 직접근거 없음"


def enrich_book(row, session):
    name = row.get("productName", "").strip()
    if not name:
        return []
    urls = [
        f"https://search.kyobobook.co.kr/search?keyword={quote_plus(name)}",
        f"https://www.yes24.com/Product/Search?domain=ALL&query={quote_plus(name)}",
        f"https://www.aladin.co.kr/search/wsearchresult.aspx?SearchTarget=All&SearchWord={quote_plus(name)}",
    ]
    evidence = []
    for url in urls:
        try:
            r = session.get(url, timeout=20)
            state = "ok" if r.ok and r.text.strip() else f"http_{r.status_code}"
            text = BeautifulSoup(r.text, "html.parser").get_text(" ", strip=True) if r.text else ""
            hit = bool(UNICORN_RE.search(text))
            evidence.append({"url": url, "state": state, "unicornHit": hit})
            if hit and row.get("publisherGrade") != "확정":
                row["publisherGrade"] = "고신뢰 추론"
                row["publisherReason"] = "공식 서점 검색 결과 유니콘 표기"
        except requests.RequestException:
            evidence.append({"url": url, "state": "request_error", "unicornHit": False})
    return evidence


def verify_kc(row, session):
    kc = (row.get("kcNumber") or "").upper()
    if not kc or kc in EXCLUDED_KC:
        row["kcNumber"] = ""
        row["kcStatus"] = "검증 불가" if row.get("responseState") != "ok" else "KC 공개표기 없음"
        row["expired"] = False
        row["noKc"] = False
        return []
    urls = [
        f"https://www.safetykorea.kr/release/certificationsearch?certNum={quote_plus(kc)}",
        f"https://www.safetykorea.kr/release/certDetail?certNum={quote_plus(kc)}",
    ]
    evidence = []
    verified = False
    expired = False
    for url in urls:
        try:
            r = session.get(url, timeout=20)
            text = BeautifulSoup(r.text, "html.parser").get_text(" ", strip=True) if r.text else ""
            contains = kc in text.upper()
            is_expired = bool(re.search(r"기간\s*만료|만료|취소|효력\s*상실", text))
            evidence.append({"url": url, "state": "ok" if r.ok else f"http_{r.status_code}", "exactNumber": contains, "expiredText": is_expired})
            verified = verified or contains
            expired = expired or (contains and is_expired)
        except requests.RequestException:
            evidence.append({"url": url, "state": "request_error", "exactNumber": False, "expiredText": False})
    row["kcStatus"] = "KC 기간만료" if expired else ("KC 적합" if verified else "검증 불가")
    row["expired"] = expired
    row["noKc"] = False
    return evidence


def write_json(path, data):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def build_excel(rows, logs, summary, out):
    wb = Workbook()
    ws = wb.active
    ws.title = "전체 유니콘 상품"
    cols = ["No.", "상품명", "쿠팡 상품 고유번호", "productId", "itemId", "vendorItemId", "상품 URL", "출판사·제조사", "출판사 판정 등급", "ISBN", "KC 인증번호", "KC 상태", "기간만료 여부", "KC 공개표기 없음 여부", "공식 No-KC 여부", "검증 근거"]
    ws.append(cols)
    selected = [r for r in rows if r.get("publisherGrade") in {"확정", "고신뢰 추론"}]
    for i, r in enumerate(selected, 1):
        ws.append([i, r.get("productName"), r.get("productId"), r.get("productId"), r.get("itemId"), r.get("vendorItemId"), r.get("productUrl"), r.get("publisherManufacturer"), r.get("publisherGrade"), r.get("isbn"), r.get("kcNumber"), r.get("kcStatus"), "Y" if r.get("expired") else "N", "Y" if r.get("kcStatus") == "KC 공개표기 없음" else "N", "Y" if r.get("noKc") else "N", r.get("publisherReason")])
    sy = wb.create_sheet("요약")
    sy.append(["항목", "값"])
    for k, v in summary.items():
        sy.append([k, v])
    lg = wb.create_sheet("검증 로그")
    lg.append(["productId", "상품명", "판정", "근거 유형", "근거 URL", "제외 사유", "확인 날짜"])
    for x in logs:
        lg.append([x.get("productId"), x.get("productName"), x.get("decision"), x.get("evidenceType"), x.get("evidenceUrl"), x.get("exclusionReason"), x.get("checkedAt")])
    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for col in range(1, sheet.max_column + 1):
            width = min(60, max(12, max(len(str(sheet.cell(r, col).value or "")) for r in range(1, sheet.max_row + 1)) + 2))
            sheet.column_dimensions[get_column_letter(col)].width = width
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--shards", type=Path, required=True)
    p.add_argument("--output-root", type=Path, default=ROOT)
    args = p.parse_args()
    products, source_states = read_shards(args.shards)
    session = requests.Session(); session.headers.update(HEADERS)
    logs = []
    for row in products:
        grade, reason = classify(row)
        row["publisherGrade"] = grade
        row["publisherReason"] = reason
        book_evidence = enrich_book(row, session) if grade == "후보" else []
        kc_evidence = verify_kc(row, session) if row["publisherGrade"] in {"확정", "고신뢰 추론"} else []
        urls = [e["url"] for e in book_evidence + kc_evidence] or row.get("sourceUrls", [])
        logs.append({"productId": row.get("productId"), "productName": row.get("productName"), "decision": row.get("publisherGrade"), "evidenceType": row.get("publisherReason"), "evidenceUrl": " | ".join(urls), "exclusionReason": "" if row.get("publisherGrade") != "제외" else row.get("publisherReason"), "checkedAt": now()})
    dup = len(products) - len({r.get("productId") for r in products})
    confirmed = [r for r in products if r.get("publisherGrade") == "확정"]
    inferred = [r for r in products if r.get("publisherGrade") == "고신뢰 추론"]
    summary = {
        "판매자 전체 상품 수": 195,
        "수집 완료 수": sum(r.get("responseState") == "ok" for r in products),
        "수집 실패 수": sum(r.get("responseState") != "ok" for r in products),
        "유니콘 확정": len(confirmed),
        "유니콘 고신뢰 추론": len(inferred),
        "KC 있음": sum(bool(r.get("kcNumber")) for r in confirmed + inferred),
        "KC 적합": sum(r.get("kcStatus") == "KC 적합" for r in confirmed + inferred),
        "KC 기간만료": sum(r.get("kcStatus") == "KC 기간만료" for r in confirmed + inferred),
        "KC 공개표기 없음": sum(r.get("kcStatus") == "KC 공개표기 없음" for r in confirmed + inferred),
        "검증 불가": sum(r.get("kcStatus") == "검증 불가" for r in confirmed + inferred),
        "중복 수": dup,
    }
    d = args.output_root / "diagnostics"
    write_json(d / "seller-products.json", products)
    write_json(d / "unresolved-products.json", [r for r in products if r.get("responseState") != "ok"])
    write_json(d / "publisher-candidates.json", [r for r in products if r.get("publisherGrade") == "후보"])
    write_json(d / "publisher-confirmed.json", confirmed + inferred)
    write_json(d / "kc-exact-mappings.json", [r for r in confirmed + inferred if r.get("kcNumber")])
    write_json(d / "final-summary.json", {"inputCount": 195, "recoveredCount": summary["수집 완료 수"], "remainingUnresolved": summary["수집 실패 수"], "newExactUnicornCount": len(confirmed), "summary": summary, "sourceStates": source_states, "generatedAt": now()})
    build_excel(products, logs, summary, args.output_root / "outputs" / "순수커머스_전체유니콘상품_KC최종.xlsx")

if __name__ == "__main__":
    main()
