#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import quote_plus

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

UNICORN_RE = re.compile(r"(?:\(주\)\s*유니콘|주식회사\s*유니콘|BOOKFRIENDS|UNICORN|(?<![가-힣])유니콘(?![가-힣]))", re.I)
KC_RE = re.compile(r"\b(?:CB|CA|SU|U)\d{3,}[A-Z0-9-]{5,}\b", re.I)
EXCLUDED_KC = {"U003E1577-7011"}
HEADERS = {"User-Agent": "Mozilla/5.0", "Accept-Language": "ko-KR,ko;q=0.9"}


def now() -> str:
    return datetime.now(timezone.utc).isoformat()


def load_products(shards: Path) -> list[dict]:
    merged: dict[str, dict] = {}
    for file in sorted(shards.rglob("*.json")):
        try:
            data = json.loads(file.read_text(encoding="utf-8"))
        except Exception:
            continue
        for row in data.get("products", []):
            pid = str(row.get("productId", ""))
            if not pid:
                continue
            cur = merged.setdefault(pid, row)
            for k, v in row.items():
                if not cur.get(k) and v:
                    cur[k] = v
            cur["sourceUrls"] = sorted(set(cur.get("sourceUrls", []) + row.get("sourceUrls", [])))
            if row.get("responseState") == "ok":
                cur["responseState"] = "ok"
    return sorted(merged.values(), key=lambda x: int(x.get("productId", 0)))


def text(url: str, session: requests.Session) -> tuple[str, str]:
    try:
        r = session.get(url, timeout=25)
        if not r.text.strip():
            return "", "empty"
        return BeautifulSoup(r.text, "html.parser").get_text(" ", strip=True), ("ok" if r.ok else f"http_{r.status_code}")
    except requests.RequestException:
        return "", "request_error"


def stage1_find_unicorn(products: list[dict], session: requests.Session, logs: list[dict]) -> list[dict]:
    out = []
    for row in products:
        direct = " ".join(str(row.get(k, "")) for k in ("productName", "publisherManufacturer", "brand", "kcText"))
        if UNICORN_RE.search(direct):
            row["publisherGrade"] = "확정"
            row["publisherReason"] = "쿠팡 상품명/필수표기 직접 표기"
            out.append(row)
            continue
        name = (row.get("productName") or "").strip()
        if not name:
            logs.append({"productId": row.get("productId"), "stage": 1, "decision": "제외", "reason": "상품명 없음", "checkedAt": now()})
            continue
        verified = False
        for base in (
            "https://search.kyobobook.co.kr/search?keyword=",
            "https://www.yes24.com/Product/Search?domain=ALL&query=",
            "https://www.aladin.co.kr/search/wsearchresult.aspx?SearchTarget=All&SearchWord=",
        ):
            url = base + quote_plus(name)
            page, state = text(url, session)
            hit = bool(UNICORN_RE.search(page))
            logs.append({"productId": row.get("productId"), "stage": 1, "decision": "공식도서검색", "reason": f"state={state}, unicornHit={hit}", "url": url, "checkedAt": now()})
            if hit:
                row["publisherGrade"] = "고신뢰 추론"
                row["publisherReason"] = "공식 도서검색 유니콘 표기"
                verified = True
                break
        if verified:
            out.append(row)
    return out


def stage2_find_kc(rows: list[dict], session: requests.Session, logs: list[dict]) -> tuple[list[dict], list[dict]]:
    missing = []
    for row in rows:
        if row.get("kcNumber") and row.get("kcNumber") not in EXCLUDED_KC:
            continue
        row["kcNumber"] = ""
        urls = [row.get("productUrl") or f"https://www.coupang.com/vp/products/{row.get('productId')}"]
        found = ""
        for url in urls:
            page, state = text(url, session)
            for candidate in KC_RE.findall(page):
                candidate = candidate.upper()
                if candidate not in EXCLUDED_KC:
                    found = candidate
                    break
            logs.append({"productId": row.get("productId"), "stage": 2, "decision": "KC상세조회", "reason": f"state={state}, found={found or 'none'}", "url": url, "checkedAt": now()})
            if found:
                break
        if found:
            row["kcNumber"] = found
            row["kcSource"] = "쿠팡 상세페이지 exact"
        else:
            missing.append(row)
    return rows, missing


def stage3_lookup_missing(rows: list[dict], missing: list[dict], evidence_path: Path, logs: list[dict]) -> None:
    evidence = json.loads(evidence_path.read_text(encoding="utf-8")) if evidence_path.exists() else []
    def norm(s: str) -> str:
        return re.sub(r"[^0-9A-Za-z가-힣]", "", s or "").lower()
    for row in missing:
        title = norm(row.get("productName", ""))
        hit = None
        for e in evidence:
            kc = str(e.get("kcNumber", "")).upper()
            model = norm(e.get("modelName", ""))
            if kc and kc not in EXCLUDED_KC and model and title and title == model:
                hit = e
                break
        if hit:
            row["kcNumber"] = hit["kcNumber"]
            row["kcStatus"] = hit.get("status", "검증 불가")
            row["kcSource"] = "기존 exact 모델명 일치"
            logs.append({"productId": row.get("productId"), "stage": 3, "decision": "기억 KC exact 매핑", "reason": hit["kcNumber"], "checkedAt": now()})
        else:
            row["kcStatus"] = "검증 불가" if row.get("responseState") != "ok" else "KC 공개표기 없음"
            row["kcSource"] = "exact 근거 없음"
            logs.append({"productId": row.get("productId"), "stage": 3, "decision": "미매핑", "reason": row["kcStatus"], "checkedAt": now()})


def verify_status(rows: list[dict], session: requests.Session, logs: list[dict]) -> None:
    for row in rows:
        kc = (row.get("kcNumber") or "").upper()
        if not kc:
            continue
        url = f"https://www.safetykorea.kr/release/certificationsearch?certNum={quote_plus(kc)}"
        page, state = text(url, session)
        exact = kc in page.upper()
        expired = bool(re.search(r"기간\s*만료|만료|취소|효력\s*상실", page)) if exact else False
        row["kcStatus"] = "KC 기간만료" if expired else ("KC 적합" if exact else "검증 불가")
        row["expired"] = expired
        logs.append({"productId": row.get("productId"), "stage": 3, "decision": "SafetyKorea 검증", "reason": f"state={state}, exact={exact}, expired={expired}", "url": url, "checkedAt": now()})


def write_excel(rows: list[dict], logs: list[dict], products: list[dict], out: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "전체 유니콘 상품"
    headers = ["No.", "상품명", "쿠팡 상품 고유번호", "productId", "itemId", "vendorItemId", "상품 URL", "출판사·제조사", "출판사 판정 등급", "ISBN", "KC 인증번호", "KC 상태", "기간만료 여부", "KC 공개표기 없음 여부", "공식 No-KC 여부", "검증 근거"]
    ws.append(headers)
    for i, r in enumerate(rows, 1):
        ws.append([i, r.get("productName"), r.get("productId"), r.get("productId"), r.get("itemId"), r.get("vendorItemId"), r.get("productUrl"), r.get("publisherManufacturer"), r.get("publisherGrade"), r.get("isbn"), r.get("kcNumber"), r.get("kcStatus"), "Y" if r.get("expired") else "N", "Y" if r.get("kcStatus") == "KC 공개표기 없음" else "N", "N", f"{r.get('publisherReason','')} | {r.get('kcSource','')}"])
    summary = {
        "판매자 전체 상품 수": 195,
        "수집 상품 수": len(products),
        "유니콘 상품 수": len(rows),
        "KC 있음": sum(bool(r.get("kcNumber")) for r in rows),
        "KC 적합": sum(r.get("kcStatus") == "KC 적합" for r in rows),
        "KC 기간만료": sum(r.get("kcStatus") == "KC 기간만료" for r in rows),
        "KC 공개표기 없음": sum(r.get("kcStatus") == "KC 공개표기 없음" for r in rows),
        "검증 불가": sum(r.get("kcStatus") == "검증 불가" for r in rows),
        "중복 수": len(products) - len({r.get("productId") for r in products}),
    }
    sy = wb.create_sheet("요약")
    sy.append(["항목", "값"])
    for k, v in summary.items(): sy.append([k, v])
    lg = wb.create_sheet("검증 로그")
    lg.append(["productId", "단계", "판정", "이유", "근거 URL", "확인 날짜"])
    for x in logs: lg.append([x.get("productId"), x.get("stage"), x.get("decision"), x.get("reason"), x.get("url"), x.get("checkedAt")])
    for sheet in wb.worksheets:
        for c in sheet[1]:
            c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="D9EAF7"); c.alignment = Alignment(horizontal="center")
        for col in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(col)].width = min(60, max(12, max(len(str(sheet.cell(r, col).value or "")) for r in range(1, sheet.max_row + 1)) + 2))
        sheet.freeze_panes = "A2"; sheet.auto_filter.ref = sheet.dimensions
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return summary


def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument("--shards", type=Path, required=True)
    p.add_argument("--root", type=Path, required=True)
    p.add_argument("--remembered-evidence", type=Path, default=Path("sellers/sunsu-commerce/diagnostics/safetykorea-status.json"))
    a = p.parse_args()
    products = load_products(a.shards)
    session = requests.Session(); session.headers.update(HEADERS)
    logs: list[dict] = []
    unicorn = stage1_find_unicorn(products, session, logs)
    unicorn, missing = stage2_find_kc(unicorn, session, logs)
    stage3_lookup_missing(unicorn, missing, a.remembered_evidence, logs)
    verify_status(unicorn, session, logs)
    d = a.root / "diagnostics"; d.mkdir(parents=True, exist_ok=True)
    (d / "seller-products.json").write_text(json.dumps(products, ensure_ascii=False, indent=2), encoding="utf-8")
    (d / "publisher-confirmed.json").write_text(json.dumps(unicorn, ensure_ascii=False, indent=2), encoding="utf-8")
    (d / "kc-missing-after-detail.json").write_text(json.dumps(missing, ensure_ascii=False, indent=2), encoding="utf-8")
    (d / "verification-log.json").write_text(json.dumps(logs, ensure_ascii=False, indent=2), encoding="utf-8")
    summary = write_excel(unicorn, logs, products, a.root / "outputs" / "순수커머스_전체유니콘상품_KC최종.xlsx")
    (d / "final-summary.json").write_text(json.dumps({"generatedAt": now(), "summary": summary}, ensure_ascii=False, indent=2), encoding="utf-8")

if __name__ == "__main__":
    main()
